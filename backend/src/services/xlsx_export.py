import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from src.api.items import ContentItemOut

HEADERS = [
    "Аккаунт",
    "Дата публикации",
    "Тип",
    "Заголовок",
    "Ссылка",
    "Описание",
    "Лайки",
    "Просмотры",
    "Дней с публикации",
    "Просмотров/день",
    "Лайков/день",
]

_TYPE_LABELS: dict[str, str] = {
    "reel": "Reels",
    "video": "Reels",
    "short": "Reels",
    "post": "Пост",
    "carousel": "Карусель",
}


def safe_filename_part(text: str) -> str:
    return re.sub(r"[^\w\-]", "_", text)[:40]


def _safe_text(value: str | None) -> str:
    """Prefix formula-trigger characters to prevent spreadsheet formula injection."""
    if value and value[0] in "=+-@":
        return f"'{value}"
    return value or ""


def build_xlsx(items: list[ContentItemOut], project_name: str, run_created_at: datetime) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)

    link_font = Font(color="0563C1", underline="single")

    for item in items:
        ws.append(
            [
                _safe_text(f"@{item.account_handle}"),
                item.published_at.replace(tzinfo=None),  # Excel doesn't handle tz-aware datetimes
                _TYPE_LABELS.get(item.type, item.type),
                _safe_text(item.title),
                item.url,  # placeholder; overwritten below as hyperlink
                _safe_text(item.summary),
                item.likes,
                item.views,
                round(item.days_since_published, 1),
                round(item.views_per_day, 1) if item.views_per_day is not None else None,
                round(item.likes_per_day, 1) if item.likes_per_day is not None else None,
            ]
        )
        row_num = ws.max_row
        url_cell = ws.cell(row=row_num, column=5)
        url_cell.hyperlink = item.url
        url_cell.font = link_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
