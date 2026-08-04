import uuid
from datetime import UTC, datetime, timedelta

from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.auth.tokens import create_access_token
from src.db import get_session
from src.main import app
from src.models import ContentType, RunStatus
from tests.conftest import (
    make_account,
    make_account_list,
    make_content_item,
    make_project,
    make_run,
    make_user,
    make_workspace,
)


class _TestClient(AsyncClient):
    async def __aexit__(self, *exc_info: object) -> None:
        app.dependency_overrides.pop(get_session, None)
        await super().__aexit__(*exc_info)


async def client(session: AsyncSession) -> _TestClient:
    async def override_get_session():
        yield session

    app.dependency_overrides[get_session] = override_get_session
    transport = ASGITransport(app=app)
    return _TestClient(transport=transport, base_url="http://test")


def auth_headers(user_id: uuid.UUID) -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token(user_id)}"}


async def _setup(session: AsyncSession):
    owner = await make_user(session)
    ws = await make_workspace(session, owner=owner)
    project = await make_project(session, workspace=ws, name="Test Project")
    account_list = await make_account_list(session, project=project)
    account = await make_account(
        session, account_list=account_list, handle="natgeo", followers_count=1000
    )
    run = await make_run(session, project=project, requested_by=owner)
    now = datetime.now(UTC)
    await make_content_item(
        session,
        run=run,
        account=account,
        type=ContentType.reel,
        published_at=now - timedelta(days=3),
        views=9000,
        likes=300,
        comments=42,
        title="Reel one",
    )
    await make_content_item(
        session,
        run=run,
        account=account,
        type=ContentType.carousel,
        published_at=now - timedelta(days=1),
        views=None,
        likes=100,
        title="Carousel two",
    )
    await session.commit()
    return owner, run


async def test_export_returns_xlsx(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    async with await client(session) as c:
        resp = await c.get(f"/runs/{run.id}/export.xlsx", headers=auth_headers(owner.id))

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = resp.headers["content-disposition"]
    assert "attachment" in cd
    assert "content-scout_" in cd
    assert ".xlsx" in cd


async def test_export_xlsx_has_russian_headers(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    async with await client(session) as c:
        resp = await c.get(f"/runs/{run.id}/export.xlsx", headers=auth_headers(owner.id))

    import io

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(1, col).value for col in range(1, 16)]
    assert headers[0] == "Аккаунт"
    assert headers[1] == "Подписчики"
    assert headers[5] == "Ссылка"
    assert headers[7] == "Лайки"
    assert headers[8] == "Просмотры"
    assert headers[9] == "Комментарии"
    assert headers[13] == "Виральность"
    assert headers[14] == "Вовлечённость"


async def test_export_xlsx_data_and_hyperlink(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    async with await client(session) as c:
        resp = await c.get(f"/runs/{run.id}/export.xlsx", headers=auth_headers(owner.id))

    import io

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws is not None

    # Two data rows + header
    assert ws.max_row == 3

    # Ссылка column (6) in row 2 should have a hyperlink
    url_cell = ws.cell(2, 6)
    assert url_cell.hyperlink is not None

    # Комментарии column (10) in row 2 (the reel with comments=42)
    assert ws.cell(2, 10).value == 42

    # Only 2 items for this account — below virality_min_items (default 3), so the badge
    # column stays blank rather than showing a misleading score off a tiny sample.
    assert ws.cell(2, 14).value in ("", None)

    # Вовлечённость (column 15): (likes + comments) / followers = (300 + 42) / 1000
    assert ws.cell(2, 15).value == (300 + 42) / 1000
    assert ws.cell(2, 15).number_format == "0.0%"


async def test_export_404_for_wrong_user(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    other = await make_user(session)
    await make_workspace(session, owner=other)
    await session.commit()
    async with await client(session) as c:
        resp = await c.get(f"/runs/{run.id}/export.xlsx", headers=auth_headers(other.id))
    assert resp.status_code == 404


async def test_export_401_without_header_or_token(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    async with await client(session) as c:
        resp = await c.get(f"/runs/{run.id}/export.xlsx")
    assert resp.status_code == 401


async def test_export_download_token_mint_and_use(session: AsyncSession) -> None:
    """Telegram's native downloadFile fetches the URL itself with no Authorization header —
    the dl_token query-param path exists specifically to cover that case."""
    owner, run = await _setup(session)
    async with await client(session) as c:
        mint_resp = await c.post(f"/runs/{run.id}/export-token", headers=auth_headers(owner.id))
        assert mint_resp.status_code == 200
        token = mint_resp.json()["token"]

        # No Authorization header at all — only the minted token.
        resp = await c.get(f"/runs/{run.id}/export.xlsx?dl_token={token}")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


async def test_export_download_token_scoped_to_its_own_run(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    other_project = await make_project(
        session, workspace=await make_workspace(session, owner=owner)
    )
    other_run = await make_run(session, project=other_project, requested_by=owner)
    await session.commit()

    async with await client(session) as c:
        mint_resp = await c.post(f"/runs/{run.id}/export-token", headers=auth_headers(owner.id))
        token = mint_resp.json()["token"]

        # Same token, different run_id in the URL — must not be accepted.
        resp = await c.get(f"/runs/{other_run.id}/export.xlsx?dl_token={token}")
        assert resp.status_code == 401


async def test_export_download_token_cannot_mint_for_foreign_run(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    other = await make_user(session)
    await make_workspace(session, owner=other)
    await session.commit()
    async with await client(session) as c:
        resp = await c.post(f"/runs/{run.id}/export-token", headers=auth_headers(other.id))
    assert resp.status_code == 404


async def test_project_items_export_token_mint_and_use(session: AsyncSession) -> None:
    owner, run = await _setup(session)
    async with await client(session) as c:
        mint_resp = await c.post(
            f"/projects/{run.project_id}/items/export-token", headers=auth_headers(owner.id)
        )
        assert mint_resp.status_code == 200
        token = mint_resp.json()["token"]

        resp = await c.get(f"/projects/{run.project_id}/items/export.xlsx?dl_token={token}")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


async def test_project_items_export_pools_across_runs_and_respects_run_filter(
    session: AsyncSession,
) -> None:
    owner = await make_user(session)
    ws = await make_workspace(session, owner=owner)
    project = await make_project(session, workspace=ws)
    account_list = await make_account_list(session, project=project)
    account = await make_account(session, account_list=account_list, handle="natgeo")
    run_1 = await make_run(session, project=project, requested_by=owner, status=RunStatus.done)
    run_2 = await make_run(session, project=project, requested_by=owner, status=RunStatus.done)
    await make_content_item(session, run=run_1, account=account, title="From run 1")
    await make_content_item(session, run=run_2, account=account, title="From run 2")
    await session.commit()

    import io

    from openpyxl import load_workbook

    async with await client(session) as c:
        headers = auth_headers(owner.id)

        resp = await c.get(f"/projects/{project.id}/items/export.xlsx", headers=headers)
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.active.max_row == 3  # header + 2 items across both runs

        resp = await c.get(
            f"/projects/{project.id}/items/export.xlsx?run_id={run_1.id}", headers=headers
        )
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.active.max_row == 2  # header + 1 item scoped to run_1


async def test_project_items_export_run_id_includes_items_from_a_failed_run(
    session: AsyncSession,
) -> None:
    """E15-S4/D43: exporting one specific (failed) run's own Publications tab reflects whatever
    it already committed — mirrors items.py:list_project_items' same fix."""
    owner = await make_user(session)
    ws = await make_workspace(session, owner=owner)
    project = await make_project(session, workspace=ws)
    account_list = await make_account_list(session, project=project)
    account = await make_account(session, account_list=account_list, handle="natgeo")
    failed_run = await make_run(
        session, project=project, requested_by=owner, status=RunStatus.failed
    )
    await make_content_item(session, run=failed_run, account=account, title="Partial item")
    await session.commit()

    import io

    from openpyxl import load_workbook

    async with await client(session) as c:
        headers = auth_headers(owner.id)

        # Aggregate export (run_id omitted) still excludes it — unchanged behavior.
        resp = await c.get(f"/projects/{project.id}/items/export.xlsx", headers=headers)
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.active.max_row == 1  # header only

        resp = await c.get(
            f"/projects/{project.id}/items/export.xlsx?run_id={failed_run.id}", headers=headers
        )
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.active.max_row == 2  # header + the one partial item


async def test_project_items_export_respects_starred_only(session: AsyncSession) -> None:
    owner = await make_user(session)
    ws = await make_workspace(session, owner=owner)
    project = await make_project(session, workspace=ws)
    account_list = await make_account_list(session, project=project)
    account = await make_account(session, account_list=account_list, handle="natgeo")
    run = await make_run(session, project=project, requested_by=owner, status=RunStatus.done)
    starred = await make_content_item(session, run=run, account=account, title="Starred")
    await make_content_item(session, run=run, account=account, title="Not starred")
    await session.commit()

    import io

    from openpyxl import load_workbook

    async with await client(session) as c:
        headers = auth_headers(owner.id)
        await c.post(
            f"/projects/{project.id}/shortlist/items",
            json={"item_ids": [str(starred.id)]},
            headers=headers,
        )

        resp = await c.get(
            f"/projects/{project.id}/items/export.xlsx?starred_only=true", headers=headers
        )
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.active.max_row == 2  # header + the one starred item


async def test_shortlist_export_download_token_mint_and_use(session: AsyncSession) -> None:
    owner = await make_user(session)
    ws = await make_workspace(session, owner=owner)
    project = await make_project(session, workspace=ws, name="Shortlist Project")
    await session.commit()

    async with await client(session) as c:
        mint_resp = await c.post(
            f"/projects/{project.id}/shortlist/export-token", headers=auth_headers(owner.id)
        )
        assert mint_resp.status_code == 200
        token = mint_resp.json()["token"]

        resp = await c.get(f"/projects/{project.id}/shortlist/export.xlsx?dl_token={token}")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
